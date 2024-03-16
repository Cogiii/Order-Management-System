from customtkinter import *
from PIL import Image
from tkcalendar import Calendar
from CTkTable import CTkTable
from datetime import datetime, timedelta
from tkinter import messagebox
import matplotlib.pyplot as plt
import openpyxl 
import os


# COLOR PALLETE FROM THIS LINK
# https://colorhunt.co/palette/5356ff378ce767c6e3dff5ff
COLOR1 = "#5356FF"
COLOR2 = "#378CE7"
COLOR3 = "#67C6E3"
COLOR4 = "#DFF5FF"

EXCEL_FILE = 'Orders.xlsx'

# Check if the Excel file exists
if os.path.exists(EXCEL_FILE):
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        sheet = workbook.active
        # Call edit method or perform other operations here
        # edit(order_id, table_number, orders, total_bill, status)
    except Exception as e:
        print(f"An error occurred while loading the workbook: {e}")
else:
    print(f"The Excel file '{EXCEL_FILE}' does not exist.")
    # Optionally, handle this case (e.g., create a new Excel file)


class OrderSystemApp():
    def __init__(self):
        self.app = CTk()
        self.app.geometry(self.center_window(self.app, 850, 700, self.app._get_window_scaling()))
        self.app.resizable(0,0)
        self.app.title("Order System")
        set_appearance_mode("light")
        self.app.iconbitmap("assets/logo.ico")

        self.no_order_id_label = None
        self.setupSidebar() # display side bar
        self.dashboardView() # display dashboard

    # Side Bar FRAME
    def setupSidebar(self):
        # Create side bar frame
        self.sidebar_frame = CTkFrame(master=self.app, fg_color=COLOR1,  width=176, height=700, corner_radius=0)
        self.sidebar_frame.pack_propagate(0)
        self.sidebar_frame.pack(fill="y", anchor="w", side="left")

        # Logo image on the top of the side bar frame
        logo_img_data = Image.open("assets/logo.png") # CHANGE LANG LOGO
        logo_img = CTkImage(dark_image=logo_img_data, light_image=logo_img_data, size=(160, 160))
        CTkLabel(master=self.sidebar_frame, text="", image=logo_img).pack(pady=(38, 0), anchor="center")

        # Dashboard Nav Bar
        analytics_img_data = Image.open("assets/analytics_icon.png")
        analytics_img = CTkImage(dark_image=analytics_img_data, light_image=analytics_img_data)
        CTkButton(master=self.sidebar_frame, image=analytics_img, text="Dashboard", fg_color="transparent", font=("Arial Bold", 14), hover_color=COLOR2, anchor="w").pack(anchor="center", ipady=5, pady=(60, 0))

        # Order Nav Bar
        package_img_data = Image.open("assets/cart_icon.png")
        package_img = CTkImage(dark_image=package_img_data, light_image=package_img_data)
        CTkButton(master=self.sidebar_frame, image=package_img, text="Orders", fg_color="transparent", font=("Arial Bold", 14), hover_color=COLOR2, anchor="w").pack(anchor="center", ipady=5, pady=(16, 0))

        # Account Informaion (Initial Only)
        person_image_data = Image.open("assets/person_icon.png")
        person_image = CTkImage(dark_image=person_image_data, light_image=person_image_data)
        CTkButton(master=self.sidebar_frame, image=person_image, text="Admin", fg_color="transparent", font=("Arial Bold", 14), hover_color=COLOR2, anchor="w").pack(anchor="center", ipady=5, pady=(300, 0))
        
        
    # Dashboard View
    def dashboardView(self):
        # Create dashboard frame
        self.dashboard_view = CTkFrame(master=self.app, fg_color="#ffffff",  width=730, height=800, corner_radius=0)
        self.dashboard_view.pack_propagate(0)
        self.dashboard_view.pack(side="left")

        # Title frame or Top Frame
        self.title_frame = CTkFrame(master=self.dashboard_view, fg_color="transparent")
        self.title_frame.pack(anchor="n", fill="x",  padx=27, pady=(29, 0))
        CTkLabel(master=self.title_frame, text="Orders", font=("Arial Black", 25), text_color=COLOR1).pack(anchor="nw", side="left")


#----------------------------------------------------------------------------------------------------------------------------------#


        # metrics order frame or status order
        self.metrics_frame = CTkFrame(master=self.dashboard_view, fg_color="transparent")
        self.metrics_frame.pack(anchor="n", fill="x",  padx=15, pady=(36, 0))

        # Total Orders Frame
        self.orders_metric = CTkFrame(master=self.metrics_frame, fg_color=COLOR1, width=200, height=80, corner_radius=12)
        self.orders_metric.grid_propagate(0)
        self.orders_metric.pack(side="left")

        # Image in Total Order Frame
        order_image_data = Image.open("assets/cart_icon.png")
        order_image = CTkImage(dark_image=order_image_data, light_image=order_image_data,size=(50, 50))
        CTkLabel(master=self.orders_metric, image=order_image, text="0").grid(row=0, column=0, rowspan=2, padx=10, pady=15)

        # Total Orders Label
        CTkLabel(master=self.orders_metric, text="Total Orders", text_color=COLOR4, font=("Arial Black", 14)).grid(row=0, column=1, sticky="sw")
        self.total_orders = CTkLabel(master=self.orders_metric, text="", text_color="#fff",font=("Arial Black", 18), justify="center")
        self.total_orders.grid(row=1, column=1, sticky="nw", pady=(0,10))


        # Total Sales metrics Frame
        self.amount_metric = CTkFrame(master=self.metrics_frame, fg_color=COLOR1, width=200, height=80, corner_radius=12)
        self.amount_metric.grid_propagate(0)        
        self.amount_metric.pack(side="left", padx=(15,0))

        # Image in Total Sales Frame
        amount_image_data = Image.open("assets/money_icon.png")
        amount_image = CTkImage(dark_image=amount_image_data, light_image=amount_image_data, size=(60, 60))
        CTkLabel(master=self.amount_metric, image=amount_image, text="").grid(row=0, column=0, rowspan=2, padx=7, pady=10)

        # Total Sales Label
        CTkLabel(master=self.amount_metric, text="Total Sales", text_color=COLOR4, font=("Arial Black", 14)).grid(row=0, column=1, sticky="sw")
        self.total_sales = CTkLabel(master=self.amount_metric, text="₱0", text_color="#fff",font=("Arial Black", 18), justify="center")
        self.total_sales.grid(row=1, column=1, sticky="nw", pady=(0,10))


        # Frame for metrics view chart and edit by order ID using grid
        self.modification_metric = CTkFrame(master=self.metrics_frame, fg_color="transparent", width=250, height=80, corner_radius=12)
        self.modification_metric.grid_propagate(0)        
        self.modification_metric.pack(side="left")

        # View Chart at grid row 0
        CTkLabel(master=self.modification_metric, text="View chart by:", text_color="black", font=("Arial Black", 12)).grid(row=0, column=0, rowspan=1, padx=0, pady=10)

        self.view_chart = CTkComboBox(master=self.modification_metric, button_color=COLOR1, border_color= COLOR2, button_hover_color=COLOR2, dropdown_hover_color=COLOR3, width=100, state="readonly", values=["Top Menu"], command=self.viewChart)
        self.view_chart.grid(row=0, column=1, rowspan=1, padx=0, pady=10)
        
        # Order Entry at grid row 1
        self.order_id = CTkEntry(master=self.modification_metric, placeholder_text="Order ID", width=100, height=25, border_width=2, border_color=COLOR1, corner_radius=10)
        self.order_id.grid(row=1,column=0, padx=10)

        CTkButton(master=self.modification_metric, text="Edit", text_color=COLOR4, font=("Arial Black",14), fg_color=COLOR1, hover_color=COLOR2, width=100, command=lambda: self.edit_order(self.order_id.get())).grid(row=1, column=1, padx=5)


# ----------------------------------------------------------------------------------------------------------------------------------- #


        # Search container or Filter container
        self.search_container = CTkFrame(master=self.dashboard_view, height=50, fg_color="#F0F0F0")
        self.search_container.pack(fill="x", pady=(30, 0), padx=27)

        # Input field for Order ID filter
        self.search_order_id = CTkEntry(master=self.search_container, width=150, placeholder_text="Search Order", border_color=COLOR3, border_width=2)
        self.search_order_id.pack(side="left", padx=(13, 0), pady=15)
        self.search_order_id.bind("<Return>", self.filter_order_id)

        current_date = datetime.now().strftime("%Y-%m-%d") # get current date
        end_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
        # Start Date Entry
        CTkLabel(master=self.search_container, text="Start Date:").pack(side="left", padx=(13, 0), pady=15)
        self.start_date = CTkEntry(master=self.search_container, width=90, border_color=COLOR3, border_width=2)
        self.start_date.pack(side="left", padx=(13, 0), pady=15)
        self.start_date.insert(0, current_date)
        self.start_date.bind("<1>", lambda event: self.pick_date(self.start_date, current_date, event))

        # End Date Entry
        CTkLabel(master=self.search_container, text="End Date:").pack(side="left", padx=(13, 0), pady=15)
        self.end_date = CTkEntry(master=self.search_container, width=90, border_color=COLOR3, border_width=2)
        self.end_date.pack(side="left", padx=(13, 0), pady=15)
        self.end_date.insert(0, end_date)
        self.end_date.bind("<1>", lambda event: (self.pick_date(self.end_date, end_date, event)))

        # Filter Button for start_date and end_date
        CTkButton(master=self.search_container,text="Filter",width=80, text_color=COLOR4, font=("Arial Black",14),fg_color=COLOR1, hover_color=COLOR2, command=self.filter_by_date).pack(side="left", padx=(13,0), pady=15)


# ----------------------------------------------------------------------------------------------------------------------------------- #

        # Initializing data for table
        self.table_header = ('Order ID','Table Number','Orders','Date','Bill','Status') # Get table header
        self.table_data = []

        
        self.table_data.append([header for header in self.table_header]) # Append all the values inside the table
        
        # Iterate over rows in descending order and append the first 12 rows
        rows_to_append = []
        start_date = int(self.start_date.get().replace("-", ""))
        end_date = int(self.end_date.get().replace("-", ""))

        for row in range(2, sheet.max_row + 1):
            values = []
            date = int(str(sheet[f"D{row}"].value).replace("-","")[:8])

            if start_date <= date <= end_date:
                rows_to_append.append(self.get_data(row))

        self.table_data.extend(reversed(rows_to_append))
        
        self.total_orders.configure(text=len(self.table_data)-1)
        self.total_sales.configure(text=f"₱{sum([float(bill[4]) for bill in self.table_data[1:]]):.2f}")

        # Create table frame 
        self.table_frame = CTkScrollableFrame(master=self.dashboard_view, fg_color="transparent",scrollbar_button_color=COLOR2, scrollbar_button_hover_color=COLOR3)
        self.table_frame.pack(expand=True, fill="both", padx=10, pady=21)
        
        # Check if table data has values
        if len(self.table_data) > 0:
            # Show tables from the table data
            self.table = CTkTable(master=self.table_frame, values=self.table_data, colors=["#E6E6E6", "#EEEEEE"], header_color=COLOR1, hover_color="#B4B4B4")
            self.table.edit_row(0, text_color="#fff", hover_color=COLOR2)
            self.table.pack(expand=True)



    # For edit order window
    def edit_order(self, order_id):
        if order_id == "":
            order_id = 0
        self.row = 0

        # get order id rows
        row_data = []
        for row in range(2,sheet.max_row + 1):
            if int(sheet[f'A{row}'].value) == int(order_id):
                self.row = row

                # 1 to 6 because there only 6 column on the table
                for col in range(1, 7):
                    if isinstance(sheet.cell(row=row, column=col).value, datetime):
                        row_data.append(sheet.cell(row=row, column=col).value.date())

                    elif col == 3:
                        orders = str(sheet.cell(row=row, column=col).value).split(", ")
                        row_data.append(orders)

                    else:
                        row_data.append(sheet.cell(row=row, column=col).value)
                    
                break        
        
        # Prompt error when row_data is empty
        if len(row_data) <= 0:
            messagebox.showerror("Error Message", "Invalid Order ID")
        else:
            # Create another window that override the order system app
            self.edit_window = CTkToplevel()
            self.edit_window.grab_set()
            self.edit_window.title("Edit")
            self.edit_window.after(250, lambda: self.edit_window.iconbitmap('assets/logo.ico'))
            self.edit_window.geometry(self.center_window(self.edit_window, 500, 350, self.edit_window._get_window_scaling()))
            self.edit_window.resizable(0, 0)

            # Display Order ID
            CTkLabel(master=self.edit_window, text=f"Order ID: {row_data[0]}", text_color=COLOR1,font=("Arial Black", 24), justify="center").place(x=10,y=10)

            # Display table number by Option
            CTkLabel(master=self.edit_window, text="Table Number:", text_color="black",font=("Arial", 15), justify="center").place(x=10,y=60)

            table_num = StringVar(value=f"{row_data[1]}")
            self.table_number = CTkComboBox(master=self.edit_window, button_color=COLOR1, border_color= COLOR2, button_hover_color=COLOR2, width=100, dropdown_hover_color=COLOR3, state="readonly", variable=table_num, 
                                            values=['1','2','3','4','5','6','7','8','9','10','TakeOut'])
            self.table_number.place(x=150,y=60)

# ----------------------------------------------------------------------------------------------------------------------------------- #

            # Display Orders
            CTkLabel(master=self.edit_window, text="Orders:", text_color="black",font=("Arial", 15), justify="center").place(x=10,y=100)
            
            # just get the value of sliders and display it dynamically
            def slider_event(value, label, item_name):
                label.configure(text=int(value))
                row_data_orders[item_name] = int(value)

            # for slider event set disabled if checkbox is off
            def orders_check(slider, quantity_label, check_var):
                slider.configure(state='disabled' if check_var.get() == 'off' else 'normal')
                if check_var.get() == "off":
                    quantity_label.configure(text=0)

            def create_components(item_name, x, y, var):
                # Check Box
                checkbox = CTkCheckBox(self.edit_window, text=item_name, width=5,
                                       command=lambda: orders_check(slider, quantity_label, var),
                                       variable=var, onvalue="on", offvalue="off")
                checkbox.place(x=x, y=y)

                quantity_label = CTkLabel(master=self.edit_window, text=f"{row_data_orders[item_name]}", font=("Arial", 15), justify="center")
                quantity_label.place(x=x+95, y=y-1)

                slider = CTkSlider(master=self.edit_window, from_=0, to=10,
                                   command=lambda value: slider_event(value, quantity_label, item_name),
                                   width=100, state='disabled' if var.get() == 'off' else 'normal')
                slider.place(x=x+115, y=y+5)

                return checkbox, quantity_label, slider

            # Check Box Variables
            pizza_var = StringVar(value="on" if "Pizza" in row_data[2] else "off")
            carbonara_var = StringVar(value="on" if "Carbonara" in row_data[2] else "off")
            chicken_var = StringVar(value="on" if "Chicken" in row_data[2] else "off")
            pancit_var = StringVar(value="on" if "Pancit" in row_data[2] else "off")
            lasagna_var = StringVar(value="on" if "Lasagna" in row_data[2] else "off")

            # Get and count the orders
            row_data_orders = {"Pizza": 0, 
                               "Carbonara": 0,
                               "Chicken": 0,
                               "Pancit": 0,
                               "Lasagna": 0}
            for order in row_data[2]:
                row_data_orders[order] += 1

            self.pizza_checkbox, self.pizza_quantity, self.pizza_slider = create_components("Pizza", 10, 130, pizza_var)
            self.carbonara_checkbox, self.carbonara_quantity, self.carbonara_slider = create_components("Carbonara", 10, 170, carbonara_var)
            self.chicken_checkbox, self.chicken_quantity, self.chicken_slider = create_components("Chicken", 10, 210, chicken_var)
            self.pancit_checkbox, self.pancit_quantity, self.pancit_slider = create_components("Pancit", 230, 130, pancit_var)
            self.lasagna_checkbox, self.lasagna_quantity, self.lasagna_slider = create_components("Lasagna", 230, 170, lasagna_var)

# ----------------------------------------------------------------------------------------------------------------------------------- #

            # Total Bill Entry
            CTkLabel(master=self.edit_window, text="Total Bill:", text_color="black",font=("Arial", 14), justify="center").place(x=10,y=250)

            self.total_bill = CTkEntry(master=self.edit_window, placeholder_text="0.00",text_color="black",font=("Arial", 13), border_color=COLOR3, height=10, fg_color="transparent")

            self.total_bill.insert(0, f"{row_data[4]:.2f}")
            self.total_bill.place(x=100, y=253)


            def display_change(event, amount, bill):
                amount, bill = float(amount), float(bill)
                change = amount - bill
                self.total_change.delete(0, END)
                self.amount_pay.delete(0, END)
                self.total_change.insert(0, f"{change:.2f}")
                self.amount_pay.insert(0, f"{amount:.2f}")

# ----------------------------------------------------------------------------------------------------------------------------------- #

            # Amount Entry
            amount_pay = StringVar(value="0.00")
            CTkLabel(master=self.edit_window, text="Amount Pay:", text_color="black",font=("Arial", 14), justify="center").place(x=10,y=280)

            self.amount_pay = CTkEntry(master=self.edit_window, placeholder_text="0",text_color="black",font=("Arial", 13), border_color=COLOR3,height=10, fg_color="transparent", textvariable=amount_pay)
            self.amount_pay.place(x=100, y=283)
            self.amount_pay.bind("<Return>", lambda event: display_change(event, self.amount_pay.get(), self.total_bill.get()))

# ----------------------------------------------------------------------------------------------------------------------------------- #

            # Change Entry
            total_change = StringVar(value="0.00")
            CTkLabel(master=self.edit_window, text="Change:", text_color="black",font=("Arial", 14), justify="center").place(x=10,y=310)

            self.total_change = CTkEntry(master=self.edit_window, placeholder_text="0",text_color="black",font=("Arial", 13),fg_color="transparent", border_color=COLOR3, height=10, textvariable=total_change)
            self.total_change.place(x=100, y=313)

# ----------------------------------------------------------------------------------------------------------------------------------- #

            # Status Entry
            status = StringVar(value=row_data[5])
            CTkLabel(master=self.edit_window, text="Status:", text_color="black",font=("Arial", 14), justify="center").place(x=265,y=253)
            self.status = CTkComboBox(master=self.edit_window, button_color=COLOR1, border_color= COLOR2, button_hover_color=COLOR2, width=100, dropdown_hover_color=COLOR3, state="readonly", values=['Serving','Pending','Dining','Completed'], variable=status)
            self.status.place(x=325,y=253)
            
# ----------------------------------------------------------------------------------------------------------------------------------- #

            # Button for edit on Edit Window
            CTkButton(master=self.edit_window, text="Edit", text_color=COLOR4, fg_color=COLOR1,font=("Arial", 15), width=90, command=lambda: self.edit_confirmation_window(order_id, row_data_orders, self.row)).place(x=300,y=300)

            # Close Button on Edit Window
            CTkButton(master=self.edit_window, text="Delete", text_color=COLOR4, fg_color="red",hover_color="#f2564b", font=("Arial", 15), width=90, command=lambda: self.delete_confirmation_window(order_id, self.row)).place(x=400,y=300)


    def delete_confirmation_window(self, order_id, row):
        self.confirmation_window = CTkToplevel()
        self.confirmation_window.grab_set()
        self.confirmation_window.title("Confirmation")
        self.confirmation_window.after(250, lambda: self.confirmation_window.iconbitmap('assets/logo.ico'))
        self.confirmation_window.geometry(self.center_window(self.confirmation_window, 260, 100, self.confirmation_window._get_window_scaling()))
        self.confirmation_window.resizable(0, 0)

        CTkLabel(master=self.confirmation_window, text=f"Are you sure you want to Delete\nOrder #{order_id} Permanently?",font=("Arial", 15)).place(x=25,y=15)

        CTkButton(master=self.confirmation_window, text="Yes", text_color=COLOR4, fg_color=COLOR1,font=("Arial", 15), width=90, command=lambda: self.delete(row)).place(x=30,y=60)

        CTkButton(master=self.confirmation_window, text="No", text_color=COLOR4, fg_color="red", hover_color="#f2564b",font=("Arial", 15), width=90, command=lambda: self.confirmation_window.destroy()).place(x=140,y=60)


    # Show window when edit is clicked for confirmation to avoid misclick
    def edit_confirmation_window(self, order_id, row_data_orders, row):
        self.confirmation_window = CTkToplevel()
        self.confirmation_window.grab_set()
        self.confirmation_window.title("Confirmation")
        self.confirmation_window.after(250, lambda: self.confirmation_window.iconbitmap('assets/logo.ico'))
        self.confirmation_window.geometry(self.center_window(self.confirmation_window, 260, 100, self.confirmation_window._get_window_scaling()))
        self.confirmation_window.resizable(0, 0)

        # Text Label Window
        CTkLabel(master=self.confirmation_window, text=f"Are you sure you want to Modify\nChanges on Order #{order_id}?",font=("Arial", 15)).place(x=25,y=15)

        # Yes button, When clicked proceed to edit function in which edit the data
        CTkButton(master=self.confirmation_window, text="Yes", text_color=COLOR4, fg_color=COLOR1,font=("Arial", 15), width=90, command=lambda: self.edit(row,
                            self.table_number.get(), 
                            row_data_orders, 
                            self.total_bill.get(), 
                            self.status.get())).place(x=30,y=60)
        
        # No button, when clicked it will just close the window
        CTkButton(master=self.confirmation_window, text="No", text_color=COLOR4, fg_color="red", hover_color="#f2564b",font=("Arial", 15), width=90, command=lambda: self.confirmation_window.destroy()).place(x=140,y=60)
    
    def delete(self, row):
        # delete the given row
        sheet.delete_rows(22, 1)
        workbook.save(EXCEL_FILE)

        # destroy the window
        self.confirmation_window.destroy()
        self.edit_window.destroy()

    # Edit the data in edit window
    def edit(self, row, table_number, orders, total_bill, status):
        print(row)
        order = ""
        
        # get the menu separated by comma to it in excel
        for menu, num in orders.items():
            order += (menu + ", ") * num

        # If table_number is a TakeOut or a string, just append it as a string else as a integer
        if table_number == "TakeOut":
            sheet[f"B{row}"] = table_number
        else:
            sheet[f"B{row}"] = int(table_number)
        
        # Change the value in excel
        sheet[f"E{row}"] = float(total_bill)
        sheet[f"F{row}"] = status
        sheet[f"C{row}"] = order[:-2] # stop at last 2 digit to erase the ", "
        workbook.save(EXCEL_FILE)

        # destroy the window
        self.confirmation_window.destroy()
        self.edit_window.destroy()
        
    # A function that view the charts 
    def viewChart(self, choice):
        start_date = self.start_date.get()
        end_date = self.end_date.get()
        filtered_start_date = int(self.start_date.get().replace("-", ""))
        filtered_end_date = int(self.end_date.get().replace("-", ""))

        # Error if start date is ahead from the end date
        if filtered_start_date > filtered_end_date:
            messagebox.showerror("Error Message", "Invalid Date")
        else:
            # If user selects menu use Pie chart to compare the top menu
            if choice == 'Top Menu':
                total_orders = {}

                # Get the row that validate the dates
                for row in range(2, sheet.max_row + 1):
                    date = int(str(sheet[f"D{row}"].value).replace("-","")[:8])

                    # Get the orders
                    if filtered_start_date <= date <= filtered_end_date:
                        orders = str(sheet[f"C{row}"].value).split(", ")

                        # Count how many orders and store in the list or dictionary
                        for order in orders:
                            if order in total_orders:
                                total_orders[order] += 1
                            else:
                                total_orders[order] = 1

                menu = [menu for menu in total_orders.keys()]
                values = [value for value in total_orders.values()]
                
                if len(menu) or len(values) > 0:
                    # Create Pie Chart
                    wp = {'linewidth': 1, 'edgecolor': COLOR1}
                    explode = [0.05] * len(values)
                    fig, ax = plt.subplots(figsize=(10, 7))
                    wedges, texts, autotexts = ax.pie(values,
                                                    autopct='%1.1f%%',
                                                    explode=explode,
                                                    labels=menu,
                                                    shadow=True,
                                                    startangle=90,
                                                    wedgeprops=wp,
                                                    textprops=dict(color="black"))
                    # Adding legend
                    # display the menu and the value
                    ax.legend(wedges, [f'{label} ({value})' for label, value in zip(menu, values)],
                            title="Menu List",
                            loc="center left",
                            bbox_to_anchor=(1, 0, 0.5, 1))
                    
                    plt.setp(autotexts, size=8, weight="bold")
                    plt.title(f"Top Menu ({start_date} - {end_date})")
                    plt.show()
                else:
                    messagebox.showerror("Error Message", "No Data Found between those dates!!!")

    # Show table filtered by date when "filter" button is click
    def filter_by_date(self):
        start_date = int(self.start_date.get().replace("-", "")) # Get the start date in numbers
        end_date = int(self.end_date.get().replace("-", "")) # Get the end date in numbers
        filtered_data = [] # data to be displayed

        filtered_data.append(self.table_header) # Append headers

        # check if start date is greater than end date then prompt error
        if start_date > end_date:
            messagebox.showerror("Error Message", "Invalid Date")
        else:
            rows_to_append = []
            
            for row in range(2, sheet.max_row + 1):
                date = int(str(sheet[f"D{row}"].value).replace("-","")[:8])

                # Check if the date is between the given date then add the row
                if start_date <= date <= end_date:
                    rows_to_append.append(self.get_data(row))
                    
            filtered_data.extend(reversed(rows_to_append)) # Set the table to descending order
        
            self.total_orders.configure(text=len(filtered_data)-1) # Get total orders
            self.total_sales.configure(text=f"₱{sum([float(bill[4]) for bill in filtered_data[1:]]):.2f}") # add all the bills
            self.update_table(filtered_data)

    def update_table(self, filtered_data):
        # Create a new table or show "Empty Table" label if no data found
        self.table.destroy()

        self.table = CTkTable(master=self.table_frame, values=filtered_data, colors=["#E6E6E6", "#EEEEEE"], header_color=COLOR1, hover_color="#B4B4B4")
        self.table.edit_row(0, text_color="#fff", hover_color=COLOR2)
        self.table.pack(expand=True)

    def get_data(self, row):
        rows_to_append = []
        count_orders = {}
        orders = []

        for col in range(1, len(self.table_header) + 1):
            # just return the value of date instead of date and time
            if isinstance(sheet.cell(row=row, column=col).value, datetime):
                rows_to_append.append(sheet.cell(row=row, column=col).value.date())
            elif col == 3:
                orders = str(sheet.cell(row=row, column=col).value).split(", ")
                for order in orders:
                    if order in count_orders:
                         count_orders[order] += 1
                    else:
                         count_orders[order] = 1
                text_order = ""
                for menu, count in count_orders.items():
                    text_order += f"{count} {menu}, "
                rows_to_append.append(text_order[:-2])
            else:
                rows_to_append.append(sheet.cell(row=row, column=col).value)
        return rows_to_append

    def filter_order_id(self, event):
        order_id = int(self.search_order_id.get())
        filtered_data = []
        row_order_ID = None

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == order_id:
                row_order_ID = i

        # check of order ID Found
        if row_order_ID:
            # Append table header
            filtered_data.append(self.table_header)
                        
            filtered_data.append(self.get_data(row_order_ID))
                
            self.update_table(filtered_data)
        else:
            messagebox.showerror("Error Message", "Invalid Order ID")


    def pick_date(self, date_entry, date,event):
        self.date_window = CTkToplevel()
        self.date_window.grab_set()
        self.date_window.title("Date")
        self.date_window.after(250, lambda: self.date_window.iconbitmap('assets/logo.ico'))
        self.date_window.geometry(self.center_window(self.date_window, 200, 200, self.date_window._get_window_scaling()))
        self.date_window.resizable(0, 0)

        cal = Calendar(self.date_window, selectmode="day", date_pattern="yyyy-mm-dd", day=int(str(date)[-2:]))
        cal.place(x=0, y=0)

        submit_btn = CTkButton(self.date_window, text="submit", fg_color=COLOR1, command=lambda: self.grab_date(date_entry, cal, self.date_window))
        submit_btn.place(x=30, y=160)


    def grab_date(self, date_entry, cal, date_window):
        selected_date = cal.get_date()
        date_entry.delete(0, END)
        date_entry.insert(0, selected_date)
        date_window.destroy()


    def center_window(self, Screen: CTk, width: int, height: int, scale_factor: float = 1.0):
        screen_width = Screen.winfo_screenwidth()
        screen_height = Screen.winfo_screenheight()
        
        x = int(((screen_width/2) - (width/2)) * scale_factor)
        y = int(((screen_height/2) - (height/2)) * scale_factor)

        return f'{width}x{height}+{x}+{y}'

    def fun(self):
        print("EXECUTED")

    def run(self):
        self.app.mainloop()

if __name__ == "__main__":
    order_system_app = OrderSystemApp()
    order_system_app.run()
